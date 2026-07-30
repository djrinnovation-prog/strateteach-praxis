"""
Telegram integration: Excel export + daily scheduler.
Sends the latest completed run's results as a multi-sheet Excel file
to a configured Telegram chat.
"""
from __future__ import annotations
import base64
import html
import io
import logging
import os
import secrets as _secrets
import threading
from datetime import datetime, timezone
from typing import Optional

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import database as db

logger = logging.getLogger(__name__)

BUCKET_COLORS = {
    "crypto":      "DBEAFE",  # blue-100
    "metals":      "FEF3C7",  # amber-100
    "stocks":      "D1FAE5",  # green-100
    "commodities": "FFEDD5",  # orange-100
}

HEADER_FILL    = PatternFill("solid", fgColor="1E40AF")   # blue-800
HEADER_FONT    = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER    = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

POS_FONT = Font(color="166534", bold=True, size=10)  # green-800
NEG_FONT = Font(color="991B1B", bold=True, size=10)  # red-800
DATA_FONT = Font(size=10)


def _pct_font(v: float) -> Font:
    return POS_FONT if v >= 0 else NEG_FONT


def build_excel(run_id: str) -> Optional[bytes]:
    """Build an Excel workbook for all results of a given run. Returns bytes."""
    results = db.get_results(run_id, sort_by="cagr", sort_dir="desc")
    if not results:
        return None

    run = db.get_run(run_id)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, run, results)

    # ── Per-bucket sheets ──────────────────────────────────────────────────────
    buckets_seen = {}
    for r in results:
        buckets_seen.setdefault(r.bucket, []).append(r)

    for bucket, bucket_results in buckets_seen.items():
        ws = wb.create_sheet(bucket.capitalize())
        _write_results_sheet(ws, bucket, bucket_results)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_summary_sheet(ws, run, results):
    from app.models import BacktestResult
    import numpy as np

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28

    title_font = Font(bold=True, size=14, color="1E40AF")
    label_font = Font(bold=True, size=10, color="374151")
    val_font   = Font(size=10, color="111827")

    ws["A1"] = "Strateteach — Backtest Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    meta = [
        ("Run Name",    run.name if run else "—"),
        ("Run ID",      run.id[:8] + "…" if run else "—"),
        ("Status",      run.status.upper() if run else "—"),
        ("Buckets",     ", ".join(run.buckets) if run else "—"),
        ("Timeframe",   run.config.timeframe if run and run.config else "—"),
        ("Date Range",  f"{run.config.startDate} → {run.config.endDate}" if run and run.config else "—"),
        ("Generated",   datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Total Symbols", len(results)),
        ("Completed",   sum(1 for r in results if not r.errorMessage)),
        ("Failed",      sum(1 for r in results if r.errorMessage)),
    ]

    ret_vals  = [r.totalReturn for r in results if not r.errorMessage]
    cagr_vals = [r.cagr       for r in results if not r.errorMessage]
    sharpe_vals = [r.sharpe   for r in results if not r.errorMessage]
    dd_vals   = [r.maxDrawdown for r in results if not r.errorMessage]

    if ret_vals:
        meta += [
            ("", ""),
            ("Avg Return %",  f"{sum(ret_vals)/len(ret_vals):.2f}%"),
            ("Avg CAGR %",    f"{sum(cagr_vals)/len(cagr_vals):.2f}%"),
            ("Avg Sharpe",    f"{sum(sharpe_vals)/len(sharpe_vals):.3f}"),
            ("Avg Max DD %",  f"{sum(dd_vals)/len(dd_vals):.2f}%"),
            ("Best Return",   f"{max(ret_vals):.2f}%"),
            ("Worst Return",  f"{min(ret_vals):.2f}%"),
        ]

    for i, (label, val) in enumerate(meta, start=3):
        if label:
            ws.cell(row=i, column=1, value=label).font = label_font
            ws.cell(row=i, column=2, value=val).font = val_font


def _write_results_sheet(ws, bucket: str, results):
    headers = ["#", "Symbol", "Name", "Total Return %", "CAGR %", "Max DD %",
               "Win Rate %", "Sharpe", "Trades", "Outlier?", "Error"]
    col_widths = [5, 16, 22, 14, 10, 11, 11, 10, 8, 10, 30]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    bucket_hex = BUCKET_COLORS.get(bucket, "F3F4F6")
    bucket_fill = PatternFill("solid", fgColor=bucket_hex)

    for row_i, r in enumerate(results, start=2):
        row_fill = bucket_fill if row_i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        outlier = "Yes" if (r.isReturnOutlier or r.isDrawdownOutlier) else ""
        vals = [
            row_i - 1, r.symbol, r.name or "", r.totalReturn, r.cagr,
            r.maxDrawdown, r.winRate, r.sharpe, r.tradeCount, outlier,
            r.errorMessage or "",
        ]
        for col_i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_i >= 4 else "left",
                                       vertical="center")
            # Colour pct columns
            if col_i in (4, 5) and isinstance(v, float):
                cell.font = _pct_font(v)
                cell.number_format = '+0.00%;-0.00%' if col_i != 6 else '0.00%'
            elif col_i == 6 and isinstance(v, float):
                cell.font = _pct_font(-v)  # drawdown: lower=worse
            else:
                cell.font = DATA_FONT


def get_latest_completed_run_id() -> Optional[str]:
    runs = db.list_runs()
    for r in runs:
        if r.status == "completed":
            return r.id
    return None


async def send_excel_to_telegram(bot_token: str, chat_id: str, run_id: Optional[str] = None) -> dict:
    """Build Excel and POST it to Telegram. Returns {ok, message}."""
    if not run_id:
        run_id = get_latest_completed_run_id()
    if not run_id:
        return {"ok": False, "message": "No completed run found to export."}

    run = db.get_run(run_id)
    xlsx_bytes = build_excel(run_id)
    if not xlsx_bytes:
        return {"ok": False, "message": "No results to export for this run."}

    filename = f"strategy_results_{datetime.now().strftime('%Y%m%d')}.xlsx"
    caption  = (
        f"Strateteach — Daily Report\n"
        f"Run: {run.name if run else run_id[:8]}\n"
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(url, data={"chat_id": chat_id, "caption": caption},
                                 files={"document": (filename, xlsx_bytes,
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

    if resp.status_code == 200 and resp.json().get("ok"):
        return {"ok": True, "message": f"Sent {filename} to Telegram."}
    else:
        err = resp.json().get("description", resp.text[:200])
        return {"ok": False, "message": f"Telegram API error: {err}"}


# ── Plain messages, group auto-detect, and event notifications ───────────────────

API = "https://api.telegram.org/bot{token}/{method}"


async def send_message(bot_token: str, chat_id: str, text: str,
                       parse_mode: Optional[str] = "HTML") -> dict:
    """POST a plain text message to a Telegram chat. Returns {ok, message}."""
    if not bot_token or not chat_id:
        return {"ok": False, "message": "Telegram not configured."}
    payload = {"chat_id": chat_id, "text": text, "disable_web_page_preview": True}
    if parse_mode:
        payload["parse_mode"] = parse_mode
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(API.format(token=bot_token, method="sendMessage"),
                                     json=payload)
        data = resp.json()
        if data.get("ok"):
            return {"ok": True, "message": "Message sent."}
        return {"ok": False, "message": data.get("description", "error")}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "message": str(exc)[:200]}


# ── Admin bot: signup approvals via inline buttons + webhook ─────────────────

def _send_buttons_sync(token: str, chat_id: str, text: str, buttons: list) -> None:
    try:
        httpx.post(API.format(token=token, method="sendMessage"),
                   json={"chat_id": chat_id, "text": text, "parse_mode": "HTML",
                         "disable_web_page_preview": True, "reply_markup": {"inline_keyboard": buttons}}, timeout=12)
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram buttons error: %s", exc)


def _send_text_sync(token: str, chat_id: str, text: str) -> None:
    try:
        httpx.post(API.format(token=token, method="sendMessage"),
                   json={"chat_id": chat_id, "text": text, "parse_mode": "HTML",
                         "disable_web_page_preview": True}, timeout=12)
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram notify_admin error: %s", exc)


def send_text_result(token: str, chat_id, text: str) -> bool:
    """Blocking send of one plain message to a chat — returns whether Telegram
    accepted it, so callers (e.g. login-code delivery) can fall back to another
    channel on failure. No-op → False if token/chat/text is missing."""
    if not token or not chat_id or not text:
        return False
    try:
        resp = httpx.post(API.format(token=token, method="sendMessage"),
                          json={"chat_id": str(chat_id), "text": text, "parse_mode": "HTML",
                                "disable_web_page_preview": True}, timeout=12)
        return bool(resp.json().get("ok"))
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram send_text_result error: %s", str(exc)[:200])
        return False


def send_to_chat(token: str, chat_id, text: str) -> None:
    """Fire-and-forget a plain message to a SPECIFIC chat (sync, non-blocking).

    Used to mirror an admin's dashboard reply back to the exact Telegram chat the
    user reached us from. No-op if token/chat is missing.
    """
    if not token or not chat_id or not text:
        return
    threading.Thread(target=_send_text_sync, args=(token, str(chat_id), text), daemon=True).start()


def send_document_sync(bot_token: str, chat_id, filename: str, data: bytes, caption: str = "",
                       mimetype: str = "application/octet-stream") -> dict:
    """Blocking sendDocument to one chat — returns {ok, message} so the caller can surface
    the real per-recipient result. Used to push the owners daily-report PDF to a connected
    owner's Telegram (the caption carries the tokenized report link)."""
    if not bot_token or not chat_id or not data:
        return {"ok": False, "message": "no telegram config / empty document"}
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        resp = httpx.post(url, data={"chat_id": str(chat_id), "caption": (caption or "")[:1000]},
                          files={"document": (filename, data, mimetype)}, timeout=30)
        try:
            body = resp.json()
        except Exception:  # noqa: BLE001
            body = {}
        if resp.status_code == 200 and body.get("ok"):
            return {"ok": True, "message": "sent"}
        return {"ok": False, "message": f"telegram: {body.get('description', resp.text[:160])}"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "message": f"telegram error: {exc}"}


def notify_admin(text: str, buttons: Optional[list] = None) -> None:
    """Fire-and-forget a message to the admin's Telegram chat (best-effort, non-blocking).

    Used to mirror app activity — suggestions, incoming chat, friend requests — to
    the admin so everything is reachable from Telegram while out.
    """
    cfg = db.get_telegram_config()
    token, chat = cfg.get("botToken"), cfg.get("chatId")
    if not token or not chat:
        return
    if buttons:
        threading.Thread(target=_send_buttons_sync, args=(token, chat, text, buttons), daemon=True).start()
    else:
        threading.Thread(target=_send_text_sync, args=(token, chat, text), daemon=True).start()


def notify_admin_sync(text: str, buttons: Optional[list] = None) -> dict:
    """Like ``notify_admin`` but SYNCHRONOUS and result-returning.

    ``notify_admin`` fires into a daemon thread and swallows every failure (a
    "not connected" config, a bad chat id, an HTTP error) as a best-effort log
    line — which is exactly why a test feedback can ping NOTHING with no trace.
    Callers that need to KNOW whether the admin alert actually reached Telegram
    (e.g. user feedback) use this instead. Returns ``{ok, reason, message}``:
      reason ∈ {sent, not_configured, api_error, exception}.
    """
    cfg = db.get_telegram_config()
    token, chat = cfg.get("botToken"), cfg.get("chatId")
    if not token or not chat:
        return {"ok": False, "reason": "not_configured",
                "message": "Admin Telegram is not connected (no bot token / chat id "
                           "saved in the Telegram screen)."}
    payload = {"chat_id": str(chat), "text": text, "parse_mode": "HTML",
               "disable_web_page_preview": True}
    if buttons:
        payload["reply_markup"] = {"inline_keyboard": buttons}
    try:
        resp = httpx.post(API.format(token=token, method="sendMessage"), json=payload, timeout=12)
        data = resp.json()
        if data.get("ok"):
            return {"ok": True, "reason": "sent", "message": "Delivered to the admin's Telegram."}
        desc = data.get("description", resp.text[:200])
        logger.warning("telegram notify_admin_sync api_error: %s", desc)
        return {"ok": False, "reason": "api_error", "message": str(desc)[:200]}
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram notify_admin_sync error: %s", str(exc)[:200])
        return {"ok": False, "reason": "exception", "message": str(exc)[:200]}


def _decode_data_url(data_url: str):
    """Split a ``data:<mime>;base64,<payload>`` URL into ``(mime, bytes)``.

    Returns ``(None, None)`` on anything that isn't a decodable base64 data URL.
    """
    try:
        head, payload = data_url.split(",", 1)
        mime = "application/octet-stream"
        if ":" in head:
            mime = head.split(":", 1)[1].split(";", 1)[0] or mime
        return mime, base64.b64decode(payload)
    except Exception:  # noqa: BLE001
        return None, None


def _send_media_sync(token: str, chat_id: str, data_url: str, caption: str,
                     filename: Optional[str]) -> None:
    mime, blob = _decode_data_url(data_url)
    if not blob:
        return
    cap = (caption or "")[:1024]
    # sendPhoto rejects files over 10MB → fall back to a document for big images.
    if mime.startswith("image/") and len(blob) <= 10_000_000:
        method, field = "sendPhoto", "photo"
    elif mime.startswith("video/"):
        method, field = "sendVideo", "video"
    else:
        method, field = "sendDocument", "document"
    ext = mime.split("/")[-1] if "/" in mime else "bin"
    name = filename or f"feedback.{ext}"
    try:
        httpx.post(API.format(token=token, method=method),
                   data={"chat_id": str(chat_id), "caption": cap},
                   files={field: (name, blob, mime or "application/octet-stream")},
                   timeout=60)
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram send_media error: %s", str(exc)[:200])


def send_media_to_admin(data_url: str, caption: str = "",
                        filename: Optional[str] = None) -> None:
    """Fire-and-forget: deliver a base64 data-URL photo/video to the admin's
    Telegram chat (sendPhoto / sendVideo / sendDocument by MIME). No-op if the
    admin Telegram isn't configured or the payload isn't a decodable data URL.
    Used by user feedback so the attached media reaches the admin immediately.
    """
    if not data_url or not isinstance(data_url, str) or "," not in data_url:
        return
    cfg = db.get_telegram_config()
    token, chat = cfg.get("botToken"), cfg.get("chatId")
    if not token or not chat:
        return
    threading.Thread(target=_send_media_sync,
                     args=(token, str(chat), data_url, caption, filename),
                     daemon=True).start()


def notify_user(username: str, text: str, buttons: Optional[list] = None) -> None:
    """Fire-and-forget a message to ONE admin's own Telegram connection.

    Used for owner-scoped alerts (e.g. a target hit on that admin's own run) so
    each admin gets their own notifications without seeing anyone else's data.
    No-op if the user hasn't connected their personal Telegram."""
    c = db.get_user_telegram(username) or {}
    token, chat = c.get("botToken"), c.get("chatId")
    if not c.get("enabled") or not token or not chat:
        return
    if buttons:
        threading.Thread(target=_send_buttons_sync, args=(token, chat, text, buttons), daemon=True).start()
    else:
        threading.Thread(target=_send_text_sync, args=(token, chat, text), daemon=True).start()


def notify_signup_request(req_id: int, name: str, contact: str, note: Optional[str] = None) -> None:
    """Push a new access request to the admin's Telegram with Approve/Deny buttons."""
    cfg = db.get_telegram_config()
    token, chat = cfg.get("botToken"), cfg.get("chatId")
    if not token or not chat:
        return
    # Sent with parse_mode=HTML — escape the visitor-supplied name/contact/note so a
    # crafted value can't inject markup (or break the Approve/Deny buttons) in the alert.
    lines = ["🆕 <b>New access request</b>", f"<b>{html.escape(name or '')}</b>", html.escape(contact or "")]
    if note:
        lines.append(f"“{html.escape(note)}”")
    text = "\n".join([l for l in lines if l])
    buttons = [[{"text": "✅ Approve", "callback_data": f"approve:{req_id}"},
                {"text": "✖ Deny", "callback_data": f"deny:{req_id}"}]]
    threading.Thread(target=_send_buttons_sync, args=(token, chat, text, buttons), daemon=True).start()


def webhook_secret() -> str:
    s = db._get_singleton("telegram_webhook_secret", None)
    if not s:
        s = _secrets.token_urlsafe(24)
        db._set_singleton("telegram_webhook_secret", s)
    return s


def app_url() -> str:
    """The public base URL for this deployment, used to build the webhook URL.

    Prefers APP_URL, then DOMAIN, then the production default. The dev override
    (docker-compose.override.yml) sets these to localhost — ``is_public_https``
    then keeps us from registering a webhook Telegram could never reach.
    """
    base = os.environ.get("APP_URL") or os.environ.get("DOMAIN") or "app.strateteach.com"
    return base if base.startswith("http") else f"https://{base}"


def is_public_https(url: str) -> bool:
    """True only for an https URL Telegram can actually reach from the internet.

    Rejects http, localhost, loopback, and *.local — i.e. the dev override — so a
    local run never clobbers production's webhook with an unreachable address.
    """
    u = (url or "").strip().lower()
    if not u.startswith("https://"):
        return False
    host = u.split("://", 1)[1].split("/", 1)[0].split("@")[-1].split(":", 1)[0]
    if not host or host in ("localhost", "127.0.0.1", "0.0.0.0", "::1"):
        return False
    if host.endswith(".local") or host.endswith(".localhost"):
        return False
    return True


async def ensure_webhook(base_url: Optional[str] = None) -> dict:
    """Idempotently (re)register the inbound webhook on startup / reconnect.

    No-op unless Telegram is connected, admin approvals are on, and the resolved
    URL is publicly reachable. Best-effort and fully logged so a misconfigured
    URL (e.g. localhost) is visible instead of silently swallowing every reply.
    """
    base = base_url or app_url()
    cfg = db.get_telegram_config()
    token, chat = cfg.get("botToken"), cfg.get("chatId")
    if not token or not chat:
        logger.info("telegram.webhook skip=not_configured")
        return {"ok": False, "message": "Telegram not configured."}
    if not db.telegram_approvals_enabled():
        logger.info("telegram.webhook skip=approvals_disabled")
        return {"ok": False, "message": "Admin approvals disabled."}
    if not is_public_https(base):
        logger.warning("telegram.webhook skip=non_public_url url=%s "
                       "(set APP_URL/DOMAIN to a public https host)", base)
        return {"ok": False, "message": f"Non-public URL — not registering: {base}"}
    hook = f"{base}/telegram/webhook"
    res = await set_webhook(token, hook, webhook_secret())
    if res.get("ok"):
        logger.info("telegram.webhook registered url=%s", hook)
    else:
        logger.warning("telegram.webhook setWebhook_failed url=%s err=%s",
                       hook, res.get("description"))
    return res


async def set_webhook(token: str, url: str, secret: str) -> dict:
    async with httpx.AsyncClient(timeout=15) as c:
        r = await c.post(API.format(token=token, method="setWebhook"),
                         json={"url": url, "secret_token": secret, "allowed_updates": ["callback_query", "message"]})
    try:
        return r.json()
    except Exception:  # noqa: BLE001
        return {"ok": False, "description": r.text[:200]}


async def delete_webhook(token: str) -> dict:
    """Remove the bot's webhook (used on Disconnect so a re-connect is clean)."""
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.post(API.format(token=token, method="deleteWebhook"),
                             json={"drop_pending_updates": True})
        return r.json()
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "description": str(e)[:200]}


async def answer_callback(token: str, cb_id: str, text: str = "") -> None:
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            await c.post(API.format(token=token, method="answerCallbackQuery"),
                         json={"callback_query_id": cb_id, "text": text})
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram answerCallbackQuery error: %s", str(exc)[:200])


async def edit_message_text(token: str, chat_id: str, message_id: int, text: str) -> None:
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            await c.post(API.format(token=token, method="editMessageText"),
                         json={"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "HTML"})
    except Exception as exc:  # noqa: BLE001
        logger.warning("telegram editMessageText error: %s", str(exc)[:200])


async def detect_chats(bot_token: str) -> dict:
    """Call getUpdates and return the distinct chats the bot can currently see.

    The bot must be a member of the group and SOMEONE must have sent a message
    there (or in a 1:1 chat with the bot) for it to appear. Returns
    {ok, chats:[{id,title,type}], message}. Newest chats first.
    """
    if not bot_token:
        return {"ok": False, "chats": [], "message": "Bot token required."}
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(API.format(token=bot_token, method="getUpdates"),
                                    params={"limit": 100, "timeout": 0})
        data = resp.json()
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "chats": [], "message": str(exc)[:200]}

    if not data.get("ok"):
        return {"ok": False, "chats": [], "message": data.get("description", "error")}

    seen: dict = {}
    # Iterate newest-first so the first time we see a chat we keep its latest title.
    for upd in reversed(data.get("result", [])):
        msg = (upd.get("message") or upd.get("channel_post")
               or upd.get("my_chat_member") or {})
        chat = msg.get("chat") or {}
        cid = chat.get("id")
        if cid is None or cid in seen:
            continue
        ctype = chat.get("type", "")
        title = (chat.get("title")
                 or " ".join(p for p in [chat.get("first_name"), chat.get("last_name")] if p)
                 or chat.get("username") or str(cid))
        seen[cid] = {"id": str(cid), "title": title, "type": ctype}

    chats = list(seen.values())
    if not chats:
        return {"ok": True, "chats": [],
                "message": ("No chats found yet. Add the bot to your group, send a "
                            "message there, then try again.")}
    return {"ok": True, "chats": chats, "message": f"Found {len(chats)} chat(s)."}


def _fmt_pct(v) -> str:
    try:
        return f"{float(v):+.2f}%"
    except (TypeError, ValueError):
        return "—"


async def notify_run_finished(run_id: str) -> dict:
    """Send a short Hebrew summary when a backtest run finishes.

    No-op (returns ok:False) unless Telegram is configured and the
    notify_run_finished toggle is on. Safe to call fire-and-forget.
    """
    cfg = db.get_telegram_config()
    if not cfg.get("botToken") or not cfg.get("chatId"):
        return {"ok": False, "message": "Not configured."}
    if not cfg.get("notifyRunFinished", True):
        return {"ok": False, "message": "Disabled."}

    run = db.get_run(run_id)
    results = db.get_results(run_id, sort_by="cagr", sort_dir="desc")
    ok_results = [r for r in results if not r.errorMessage]
    n_total = len(results)
    n_ok = len(ok_results)
    n_fail = n_total - n_ok

    avg_ret = (sum(r.totalReturn for r in ok_results) / n_ok) if n_ok else 0.0
    avg_cagr = (sum(r.cagr for r in ok_results) / n_ok) if n_ok else 0.0
    top = ok_results[0] if ok_results else None

    lines = [
        "✅ <b>ריצת בקטסט הסתיימה</b>",
        f"שם: <b>{run.name if run else run_id[:8]}</b>",
        f"נכסים: {n_total} | הצליחו: {n_ok} | נכשלו: {n_fail}",
    ]
    if n_ok:
        lines.append(f"תשואה ממוצעת: <b>{_fmt_pct(avg_ret)}</b> · CAGR ממוצע: <b>{_fmt_pct(avg_cagr)}</b>")
    if top:
        lines.append(f"מוביל: <b>{top.symbol}</b> ({_fmt_pct(top.totalReturn)})")
    lines.append(f"\n<i>{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</i>")

    result = await send_message(cfg["botToken"], cfg["chatId"], "\n".join(lines))
    db.update_telegram_send_status(result["ok"], result.get("message", ""))
    return result


async def notify_profit_plan(plan: dict) -> dict:
    """Send a Hebrew Profit Engine alert for a freshly built plan."""
    cfg = db.get_telegram_config()
    if not cfg.get("botToken") or not cfg.get("chatId"):
        return {"ok": False, "message": "Not configured."}
    if not cfg.get("notifyProfitEngine", True):
        return {"ok": False, "message": "Disabled."}

    items = plan.get("items") or []
    if not items:
        return {"ok": False, "message": "Empty plan — nothing to send."}

    quote = plan.get("quote", "USDT")
    lines = [
        "💡 <b>מנוע המסחר — תוכנית חדשה</b>",
        f"יעד {('יומי' if plan.get('mode') == 'daily' else 'שבועי')}: "
        f"<b>{plan.get('target', 0)} {quote}</b> · "
        f"להשקעה: <b>{plan.get('deployable', 0)} {quote}</b>",
        f"מספר עסקאות מוצע: <b>{plan.get('nTrades', 0)}</b>",
        "",
    ]
    for it in items[:10]:
        lines.append(
            f"• <b>{it.get('symbol')}</b> — קנייה ב-{it.get('entryPrice')} → "
            f"מכירה ב-{it.get('takeProfitPrice')} "
            f"(תנועה נדרשת {_fmt_pct(it.get('requiredMovePct'))})"
        )
    lines.append("\n<i>המלצה בלבד — כל פקודה מאושרת ידנית עם קוד אישי באפליקציה.</i>")

    result = await send_message(cfg["botToken"], cfg["chatId"], "\n".join(lines))
    db.update_telegram_send_status(result["ok"], result.get("message", ""))
    return result


_TIER_EMOJI = {"breaking_out": "🚀", "near_breakout": "⚡",
               "in_uptrend": "✅", "building": "📈"}


async def notify_signals(signals: list) -> dict:
    """Send a Hebrew daily signal alert (top breakout/trend picks)."""
    cfg = db.get_telegram_config()
    if not cfg.get("botToken") or not cfg.get("chatId"):
        return {"ok": False, "message": "Not configured."}
    if not cfg.get("notifySignals", True):
        return {"ok": False, "message": "Disabled."}

    sigs = signals or []
    if not sigs:
        return {"ok": False, "message": "No signals to send."}

    lines = ["📊 <b>סיגנלים להיום — פריצות ומגמות</b>", ""]
    for s in sigs[:12]:
        em = _TIER_EMOJI.get(s.get("tier", ""), "•")
        name = s.get("name") or s.get("symbol") or "?"
        bucket = s.get("bucket", "")
        desc = s.get("desc", "")
        lines.append(f"{em} <b>{name}</b> ({bucket}) — {desc}")
    lines.append(f"\n<i>{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</i>")

    result = await send_message(cfg["botToken"], cfg["chatId"], "\n".join(lines))
    db.update_telegram_send_status(result["ok"], result.get("message", ""))
    return result
